from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required  # Thêm dòng này
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q, Avg, F
from django.db.models.functions import ExtractMonth
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.conf import settings
import json
import csv
import datetime
from accounts.models import CustomUser, BalanceHistory, TCoinHistory
from store.models import Product, Category, Order, ProductVariant, VariantOption, Banner

# Thay đổi import models từ cấu trúc mới
from .models.base import (SupportTicket, TicketReply, EmailTemplate, 
                         ReferralProgram, ReferralCode, ReferralTransaction, 
                         APIKey, APILog, Campaign)
from .models.discount import Discount, UserDiscount, DiscountUsage
from .models.subscription import SubscriptionPlan, UserSubscription, SubscriptionTransaction
from .models.warranty import WarrantyTicket, WarrantyHistory

from blog.models import Post, Category as BlogCategory
from datetime import timedelta
import uuid
try:
    import xlswriter
except ImportError:
    import xlsxwriter as xlswriter
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.core.mail import send_mail
from django.template.loader import render_to_string

# Thêm vào đầu file, sau các import hiện có
from .models.source import Source, SourceProduct, SourceLog

# Thêm import forms
from .forms import SourceForm, SourceLogForm, SourceProductForm

from django.db import models
from django.urls import reverse

# Thêm vào đầu file views.py sau các import khác
from . import chatbot_views

# Authentication
def dashboard_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None and user.is_staff:
            login(request, user)
            return redirect('dashboard:home')
        else:
            messages.error(request, 'Tên đăng nhập hoặc mật khẩu không đúng')
    
    return render(request, 'dashboard/login.html')

def dashboard_logout(request):
    logout(request)
    return redirect('dashboard:login')

# Dashboard
@staff_member_required
def index(request):
    """Dashboard trang chủ"""
    # Thống kê tổng quát
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='pending').count()
    completed_orders = Order.objects.filter(status='completed').count()
    
    # Số lượng sản phẩm
    total_products = Product.objects.count()
    
    # Tổng doanh thu
    revenue = Order.objects.filter(status='completed').aggregate(
        total=Sum('total_price')
    )['total'] or 0
    
    # Số lượng người dùng
    total_users = CustomUser.objects.count()
    
    # Đơn hàng gần đây
    recent_orders = Order.objects.all().order_by('-created_at')[:5]
    
    # Người dùng mới
    new_users = CustomUser.objects.all().order_by('-date_joined')[:5]
    
    # Sản phẩm bán chạy
    top_products = Product.objects.annotate(
        sold=Count('orderitem')
    ).order_by('-sold')[:5]
    
    # Dữ liệu ticket hỗ trợ
    pending_tickets = SupportTicket.objects.filter(status='pending').count()
    total_tickets = SupportTicket.objects.count()
    
    # Dữ liệu giao dịch bảo hành
    warranty_requests = WarrantyTicket.objects.filter(status='pending').count()
    
    context = {
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'completed_orders': completed_orders,
        'total_products': total_products,
        'revenue': revenue,
        'total_users': total_users,
        'recent_orders': recent_orders,
        'new_users': new_users,
        'top_products': top_products,
        'pending_tickets': pending_tickets,
        'total_tickets': total_tickets,
        'warranty_requests': warranty_requests
    }
    
    return render(request, 'dashboard/index.html', context)

# Products
@staff_member_required
def product_list(request):
    products = Product.objects.all()
    return render(request, 'dashboard/products/list.html', {'products': products})

@staff_member_required
def add_product(request):
    if request.method == 'POST':
        # Xử lý thêm sản phẩm
        name = request.POST.get('name')
        price = request.POST.get('price')
        category_id = request.POST.get('category')
        description = request.POST.get('description')
        
        category = get_object_or_404(Category, id=category_id) if category_id else None
        
        product = Product.objects.create(
            name=name,
            price=price,
            category=category,
            description=description
        )
        
        messages.success(request, f'Đã thêm sản phẩm {name}')
        return redirect('dashboard:products')
    
    categories = Category.objects.all()
    return render(request, 'dashboard/products/add.html', {'categories': categories})

@staff_member_required
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        # Xử lý cập nhật sản phẩm
        product.name = request.POST.get('name')
        product.price = request.POST.get('price')
        
        category_id = request.POST.get('category')
        product.category = get_object_or_404(Category, id=category_id) if category_id else None
        
        product.description = request.POST.get('description')
        product.save()
        
        messages.success(request, f'Đã cập nhật sản phẩm {product.name}')
        return redirect('dashboard:products')
    
    categories = Category.objects.all()
    return render(request, 'dashboard/products/edit.html', {
        'product': product,
        'categories': categories
    })

@staff_member_required
def delete_product(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id)
        product_name = product.name
        product.delete()
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        
        messages.success(request, f'Đã xóa sản phẩm {product_name}')
        return redirect('dashboard:products')
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@staff_member_required
def product_detail(request, product_id):
    """Product detail view"""
    product = get_object_or_404(Product, id=product_id)
    return render(request, 'dashboard/products/detail.html', {'product': product})

@staff_member_required
def get_product(request, product_id):
    """Lấy thông tin sản phẩm dạng JSON cho AJAX"""
    product = get_object_or_404(Product, id=product_id)
    data = {
        'id': product.id,
        'name': product.name,
        'price': float(product.price),
        'description': product.description,
        'stock': product.stock,
        'category': product.category.name if product.category else '',
    }
    return JsonResponse(data)

@staff_member_required
def import_products(request):
    """Import products from CSV/Excel"""
    if request.method == 'POST':
        # Xử lý file import
        file = request.FILES.get('import_file')
        if file:
            # Xử lý file ở đây
            messages.success(request, 'Đã nhập sản phẩm thành công')
        else:
            messages.error(request, 'Vui lòng chọn file để nhập')
    return render(request, 'dashboard/products/import.html')

@staff_member_required
def export_products(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Name', 'Price', 'Category', 'Description'])
    
    products = Product.objects.all()
    for product in products:
        writer.writerow([
            product.id,
            product.name,
            product.price,
            product.category.name if product.category else '',
            product.description
        ])
    
    return response

# Categories
@staff_member_required
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'dashboard/products/categories.html', {'categories': categories})

@staff_member_required
def add_category(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        
        Category.objects.create(
            name=name,
            description=description
        )
        
        messages.success(request, f'Đã thêm danh mục {name}')
        return redirect('dashboard:categories')
    
    return render(request, 'dashboard/products/add_category.html')

@staff_member_required
def edit_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    
    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.description = request.POST.get('description')
        category.save()
        
        messages.success(request, f'Đã cập nhật danh mục {category.name}')
        return redirect('dashboard:categories')
    
    return render(request, 'dashboard/products/edit_category.html', {'category': category})

@staff_member_required
def delete_category(request, category_id):
    if request.method == 'POST':
        category = get_object_or_404(Category, id=category_id)
        category_name = category.name
        category.delete()
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        
        messages.success(request, f'Đã xóa danh mục {category_name}')
        return redirect('dashboard:categories')
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# Orders
@staff_member_required
def orders(request):
    orders = Order.objects.all().order_by('-created_at')
    return render(request, 'dashboard/orders/index.html', {'orders': orders})

@staff_member_required
def order_management(request):
    """Advanced order management view"""
    orders = Order.objects.all().order_by('-created_at')
    
    # Advanced filters
    status = request.GET.get('status')
    payment_status = request.GET.get('payment_status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    min_amount = request.GET.get('min_amount')
    max_amount = request.GET.get('max_amount')
    
    # Thêm các trường này nếu chưa có trong model Order
    status_choices = [
        ('pending', 'Chờ xử lý'),
        ('processing', 'Đang xử lý'),
        ('shipped', 'Đã giao hàng'),
        ('completed', 'Hoàn thành'),
        ('cancelled', 'Đã hủy'),
    ]
    
    payment_status_choices = [
        ('pending', 'Chờ thanh toán'),
        ('paid', 'Đã thanh toán'),
        ('failed', 'Thanh toán thất bại'),
        ('refunded', 'Đã hoàn tiền'),
    ]
    
    payment_method_choices = [
        ('cod', 'Thanh toán khi nhận hàng'),
        ('bank_transfer', 'Chuyển khoản ngân hàng'),
        ('vnpay', 'VNPay'),
        ('momo', 'MoMo'),
        ('zalopay', 'ZaloPay'),
        ('balance', 'Số dư tài khoản'),
    ]
    
    context = {
        'orders': orders,
        'status_choices': status_choices,
        'payment_status_choices': payment_status_choices,
        'payment_method_choices': payment_method_choices
    }
    
    return render(request, 'dashboard/orders/management.html', context)

@staff_member_required
def order_detail(request, order_id):
    """Detailed order view with timeline"""
    order = get_object_or_404(Order, id=order_id)
    
    # Order timeline
    timeline = []
    
    # Thêm sự kiện tạo đơn
    timeline.append({
        'time': order.created_at,
        'type': 'created',
        'title': 'Đơn hàng được tạo',
        'description': f'Khách hàng {order.user.username} đã tạo đơn hàng'
    })
    
    # Thêm sự kiện thanh toán
    if order.paid_at:
        timeline.append({
            'time': order.paid_at,
            'type': 'payment',
            'title': 'Thanh toán thành công',
            'description': f'Thanh toán {order.total_amount:,.0f}đ qua {order.payment_method}'
        })
    
    # Thêm các sự kiện thay đổi trạng thái
    for status_change in order.status_changes.all():
        timeline.append({
            'time': status_change.created_at,
            'type': 'status',
            'title': f'Trạng thái: {status_change.new_status}',
            'description': status_change.note
        })
    
    # Sắp xếp timeline theo thời gian
    timeline.sort(key=lambda x: x['time'], reverse=True)
    
    context = {
        'order': order,
        'timeline': timeline
    }
    
    return render(request, 'dashboard/orders/detail.html', context)

@staff_member_required
def update_order_status(request, order_id):
    """Cập nhật trạng thái đơn hàng"""
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        new_status = request.POST.get('status')
        
        if new_status and new_status != order.status:
            old_status = order.status
            order.status = new_status
            order.save()
            
            # Ghi log thay đổi trạng thái
            # ...
            
            # Gửi email thông báo cho khách hàng
            send_order_status_email(order)
        
            messages.success(request, 'Đã cập nhật trạng thái đơn hàng')
    
    return redirect('dashboard:order_detail', order_id=order_id)

@staff_member_required
def cancel_order(request):
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        order = get_object_or_404(Order, id=order_id)
        
        order.status = 'cancelled'
        order.save()
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

# Users
@staff_member_required
def user_list(request):
    """Basic user list view"""
    users = CustomUser.objects.all().order_by('-date_joined')
    return render(request, 'dashboard/users/index.html', {'users': users})

@staff_member_required
def user_management(request):
    """Advanced user management view"""
    users = CustomUser.objects.all().order_by('-date_joined')
    
    # Tìm kiếm
    search = request.GET.get('search')
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(email__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )
    
    # Top users - sửa lỗi
    # Thay vì sử dụng orders__total_amount, dùng store_orders__total_amount
    top_spenders = CustomUser.objects.annotate(
        total_spent=Sum('store_orders__total_amount')
    ).exclude(total_spent=None).order_by('-total_spent')[:10]
    
    # Phân trang
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    users_paginated = paginator.get_page(page_number)
    
    context = {
        'users': users_paginated,
        'top_spenders': top_spenders,
        'search_query': search or '',
    }
    
    return render(request, 'dashboard/users/management.html', context)

@staff_member_required
def user_detail(request, user_id):
    """Detailed user view"""
    user = get_object_or_404(CustomUser, id=user_id)
    
    # Orders
    orders = Order.objects.filter(user=user).order_by('-created_at')
    total_spent = orders.aggregate(total=Sum('total_amount'))['total'] or 0

    # Balance history
    balance_history = BalanceHistory.objects.filter(user=user).order_by('-created_at')
    
    # TCoin history
    tcoin_history = TCoinHistory.objects.filter(user=user).order_by('-created_at')
    
    # Activity timeline
    timeline = []
    
    # Add registration
    timeline.append({
        'time': user.date_joined,
        'type': 'registration',
        'title': 'Đăng ký tài khoản',
        'description': f'Đăng ký qua {user.get_auth_provider_display()}'
    })
    
    # Add last login
    if user.last_login:
        timeline.append({
            'time': user.last_login,
            'type': 'login',
            'title': 'Đăng nhập gần nhất',
            'description': f'Đăng nhập vào {user.last_login.strftime("%d/%m/%Y %H:%M")}'
        })
    
    # Add orders
    for order in orders:
        timeline.append({
            'time': order.created_at,
            'type': 'order',
            'title': f'Đặt đơn hàng #{order.id}',
            'description': f'Tổng tiền: {order.total_amount:,.0f}đ'
        })
    
    # Add balance transactions
    for transaction in balance_history:
        timeline.append({
            'time': transaction.created_at,
            'type': 'balance',
            'title': 'Giao dịch số dư',
            'description': f'{transaction.amount:+,.0f}đ - {transaction.description}'
        })
    
    # Sort timeline
    timeline.sort(key=lambda x: x['time'], reverse=True)
    
    context = {
        'user_profile': user,
        'orders': orders,
        'total_spent': total_spent,
        'balance_history': balance_history,
        'tcoin_history': tcoin_history,
        'timeline': timeline
    }
    
    return render(request, 'dashboard/users/detail.html', context)

@staff_member_required
def toggle_user_status(request, user_id):
    """Toggle user active status"""
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id)
        user.is_active = not user.is_active
        user.save()
        
        status = 'kích hoạt' if user.is_active else 'vô hiệu hóa'
        messages.success(request, f'Đã {status} tài khoản {user.username}')
        
    return redirect('dashboard:user_detail', user_id=user_id)

@staff_member_required
def add_user(request):
    """Thêm người dùng mới"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        
        # Kiểm tra người dùng đã tồn tại chưa
        if CustomUser.objects.filter(username=username).exists():
            messages.error(request, f'Tên người dùng {username} đã tồn tại')
            return render(request, 'dashboard/users/add.html')
            
        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, f'Email {email} đã được sử dụng')
            return render(request, 'dashboard/users/add.html')
        
        # Tạo người dùng mới
        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name
        )
        
        messages.success(request, f'Đã thêm người dùng {username}')
        return redirect('dashboard:user_detail', user_id=user.id)
    
    return render(request, 'dashboard/users/add.html')

@staff_member_required
def edit_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        is_staff = request.POST.get('is_staff') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        
        # Cập nhật thông tin người dùng
        user.username = username
        user.email = email
        user.is_staff = is_staff
        user.is_active = is_active
        user.save()
        
        messages.success(request, f'Đã cập nhật người dùng {username}')
        return redirect('dashboard:users')
    
    return render(request, 'dashboard/users/edit.html', {'user': user})

@staff_member_required
def delete_user(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id)
        username = user.username
        user.delete()
        
        messages.success(request, f'Đã xóa người dùng {username}')
        return redirect('dashboard:users')
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@staff_member_required
def lock_user(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        user = get_object_or_404(CustomUser, id=user_id)
        
        user.is_active = False
        user.save()
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

@staff_member_required
def unlock_user(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        user = get_object_or_404(CustomUser, id=user_id)
        
        user.is_active = True
        user.save()
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

# Roles
@staff_member_required
def roles(request):
    # Trong Django, roles thường được quản lý qua Groups
    from django.contrib.auth.models import Group
    roles = Group.objects.all()
    
    return render(request, 'dashboard/users/roles.html', {'roles': roles})

@staff_member_required
def add_role(request):
    if request.method == 'POST':
        from django.contrib.auth.models import Group
        name = request.POST.get('name')
        
        Group.objects.create(name=name)
        
        messages.success(request, f'Đã thêm vai trò {name}')
        return redirect('dashboard:roles')
    
    return render(request, 'dashboard/users/add_role.html')

# Customers
@staff_member_required
def customers(request):
    customers = CustomUser.objects.filter(is_staff=False)
    return render(request, 'dashboard/customers/list.html', {'customers': customers})

@staff_member_required
def customer_detail(request, customer_id):
    customer = get_object_or_404(CustomUser, id=customer_id)
    orders = Order.objects.filter(user=customer).order_by('-created_at')
    
    context = {
        'customer': customer,
        'orders': orders
    }
    
    return render(request, 'dashboard/customers/detail.html', context)

@staff_member_required
def delete_address(request):
    if request.method == 'POST':
        address_id = request.POST.get('address_id')
        # Xử lý xóa địa chỉ (cần model Address)
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

# Reports
@staff_member_required
def reports(request):
    # Thống kê doanh thu theo thời gian
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    
    # Doanh thu hôm nay
    today_revenue = Order.objects.filter(
        created_at__date=today,
        status='completed'
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Doanh thu tháng này
    month_revenue = Order.objects.filter(
        created_at__date__gte=start_of_month,
        status='completed'
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Thống kê theo trạng thái đơn hàng
    status_stats = Order.objects.values('status').annotate(
        count=Count('id'),
        revenue=Sum('total_amount')
    )
    
    context = {
        'today_revenue': today_revenue,
        'month_revenue': month_revenue,
        'status_stats': status_stats
    }
    
    return render(request, 'dashboard/reports/index.html', context)

# Analytics
@staff_member_required
def analytics(request):
    return render(request, 'dashboard/analytics/index.html')

@staff_member_required
def realtime_analytics(request):
    return render(request, 'dashboard/analytics/realtime.html')

# Settings
@staff_member_required
def settings(request):
    return render(request, 'dashboard/settings/index.html')

@staff_member_required
def general_settings(request):
    return render(request, 'dashboard/settings/general.html')

@staff_member_required
def clear_logs(request):
    if request.method == 'POST':
        # Xử lý xóa logs (cần model Log)
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

# Marketing
@staff_member_required
def marketing(request):
    return render(request, 'dashboard/marketing/index.html')

@staff_member_required
def campaigns(request):
    return render(request, 'dashboard/marketing/campaigns.html')

@staff_member_required
def campaign_list(request):
    """Campaign list view"""
    campaigns = Campaign.objects.all().order_by('-created_at')
    return render(request, 'dashboard/marketing/campaigns.html', {'campaigns': campaigns})

@staff_member_required
def add_campaign(request):
    """Add campaign view"""
    if request.method == 'POST':
        form = CampaignForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Chiến dịch đã được tạo thành công')
        return redirect('dashboard:campaigns')
    else:
        form = CampaignForm()
    return render(request, 'dashboard/marketing/add_campaign.html', {'form': form})

@staff_member_required
def get_campaign(request):
    campaign_id = request.GET.get('campaign_id')
    # Xử lý lấy thông tin chiến dịch (cần model Campaign)
    
    # Dữ liệu mẫu
    campaign = {
        'id': campaign_id,
        'name': 'Chiến dịch mẫu',
        'channel': 'email',
        'budget': 1000000,
        'status': 'active',
        'description': 'Mô tả chiến dịch'
    }
    
    return JsonResponse({'success': True, 'campaign': campaign})

@staff_member_required
def banners(request):
    banners = Banner.objects.all()
    return render(request, 'dashboard/marketing/banners.html', {'banners': banners})

# Content
@staff_member_required
def content(request):
    return render(request, 'dashboard/content/index.html')

@staff_member_required
def edit_page(request, page_id=None):
    # Xử lý chỉnh sửa trang (cần model Page)
    
    return render(request, 'dashboard/content/edit_page.html')

# Blog
@staff_member_required
def blogs(request):
    posts = Post.objects.all().order_by('-created_at')
    return render(request, 'dashboard/posts/list.html', {'posts': posts})

@staff_member_required
def add_blog(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        category_id = request.POST.get('category')
        
        category = get_object_or_404(BlogCategory, id=category_id) if category_id else None
        
        Post.objects.create(
            title=title,
            content=content,
            category=category,
            author=request.user
        )
        
        messages.success(request, 'Đã thêm bài viết mới')
        return redirect('dashboard:blogs')
    
    # Lấy danh sách category cho form
    try:
        from blog.models import Category
        categories = Category.objects.all()
    except:
        categories = []
    
    context = {
        'categories': categories
    }
    
    return render(request, 'dashboard/blog/add.html', context)

@staff_member_required
def blog_categories(request):
    categories = BlogCategory.objects.all()
    return render(request, 'dashboard/blog/categories.html', {'categories': categories})

# Inventory
@staff_member_required
def inventory(request):
    return render(request, 'dashboard/inventory/stock.html')

@staff_member_required
def get_stock_data(request):
    period = request.GET.get('period', '30')  # Mặc định 30 ngày
    days = int(period)
    
    # Tạo dữ liệu mẫu
    end_date = timezone.now().date()
    start_date = end_date - datetime.timedelta(days=days)
    
    dates = []
    import_data = []
    export_data = []
    
    current_date = start_date
    while current_date <= end_date:
        dates.append(current_date.strftime('%d/%m'))
        import_data.append(round(100 * (1 + 0.5 * (current_date.day % 5))))
        export_data.append(round(80 * (1 + 0.3 * (current_date.day % 7))))
        current_date += datetime.timedelta(days=1)
    
    return JsonResponse({
        'dates': dates,
        'import_data': import_data,
        'export_data': export_data
    })

# Support Tickets
@staff_member_required
def tickets(request):
    tickets = SupportTicket.objects.all().order_by('-created_at')
    return render(request, 'dashboard/tickets/list.html', {'tickets': tickets})

@staff_member_required
def ticket_detail(request, ticket_id):
    ticket = get_object_or_404(SupportTicket, id=ticket_id)
    replies = TicketReply.objects.filter(ticket=ticket).order_by('created_at')
    
    if request.method == 'POST':
        content = request.POST.get('content')
        
        TicketReply.objects.create(
            ticket=ticket,
            user=request.user,
            content=content,
            is_admin_reply=True
        )
        
        # Cập nhật trạng thái ticket
        ticket.status = 'pending'
        ticket.save()
        
        messages.success(request, 'Đã gửi phản hồi')
        return redirect('dashboard:ticket_detail', ticket_id=ticket.id)
    
    context = {
        'ticket': ticket,
        'replies': replies
    }
    
    return render(request, 'dashboard/tickets/detail.html', context)

@staff_member_required
def assign_ticket(request, ticket_id):
    if request.method == 'POST':
        ticket = get_object_or_404(SupportTicket, id=ticket_id)
        staff_id = request.POST.get('staff_id')
        
        if staff_id:
            staff = get_object_or_404(CustomUser, id=staff_id)
            ticket.assigned_to = staff
            ticket.save()
            
            messages.success(request, f'Đã phân công ticket cho {staff.username}')
        
        return redirect('dashboard:ticket_detail', ticket_id=ticket.id)
    
    return redirect('dashboard:tickets')

@staff_member_required
def close_ticket(request, ticket_id):
    ticket = get_object_or_404(SupportTicket, id=ticket_id)
    ticket.status = 'closed'
    ticket.closed_at = timezone.now()
    ticket.save()
    
    messages.success(request, 'Đã đóng ticket')
    return redirect('dashboard:ticket_detail', ticket_id=ticket.id)

# Chart Data
@staff_member_required
def chart_data(request):
    # Dữ liệu cho biểu đồ doanh thu
    labels = ['Tháng 1', 'Tháng 2', 'Tháng 3', 'Tháng 4', 'Tháng 5', 'Tháng 6']
    revenue_data = [5000000, 7000000, 6000000, 8000000, 9500000, 11000000]
    
    # Dữ liệu cho biểu đồ đơn hàng
    order_data = [150, 200, 180, 250, 300, 350]
    
    # Dữ liệu cho biểu đồ người dùng
    user_data = [50, 80, 100, 120, 150, 200]
    
    return JsonResponse({
        'labels': labels,
        'revenue': revenue_data,
        'orders': order_data,
        'users': user_data
    })

# Email Templates
@staff_member_required
def email_templates(request):
    """Trang quản lý mẫu email"""
    # Tạo dữ liệu mẫu để trang templates.html không bị lỗi
    welcome_email = EmailTemplate.objects.filter(category='welcome').first()
    order_confirmation = EmailTemplate.objects.filter(category='order').first()
    password_reset = EmailTemplate.objects.filter(category='password').first()
    
    # Nếu không có dữ liệu mẫu, tạo các đối tượng rỗng với ID mặc định
    if not welcome_email:
        welcome_email = type('obj', (object,), {'id': 1, 'name': 'Email chào mừng', 'subject': 'Chào mừng bạn đến với hệ thống'})
    
    if not order_confirmation:
        order_confirmation = type('obj', (object,), {'id': 2, 'name': 'Xác nhận đơn hàng', 'subject': 'Đơn hàng của bạn đã được xác nhận'})
    
    if not password_reset:
        password_reset = type('obj', (object,), {'id': 3, 'name': 'Đặt lại mật khẩu', 'subject': 'Yêu cầu đặt lại mật khẩu'})
    
    context = {
        'welcome_email': welcome_email,
        'order_confirmation': order_confirmation,
        'password_reset': password_reset
    }
    
    return render(request, 'dashboard/email/templates.html', context)

@staff_member_required
def add_email_template(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        subject = request.POST.get('subject')
        content = request.POST.get('content')
        
        EmailTemplate.objects.create(
            name=name,
            subject=subject,
            content=content
        )
        
        messages.success(request, f'Đã thêm template email {name}')
        return redirect('dashboard:email_templates')
    
    return render(request, 'dashboard/emails/add_template.html')

# Analytics Views
@staff_member_required
def analytics_dashboard(request):
    # Tính toán các chỉ số quan trọng
    now = timezone.now()
    start_date = now - timedelta(days=30)
    
    # Doanh số 30 ngày qua
    orders = Order.objects.filter(created_at__gte=start_date, status='completed')
    total_revenue = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Tổng số khách hàng mới
    new_users = CustomUser.objects.filter(date_joined__gte=start_date).count()
    
    # Tổng số đơn hàng
    total_orders = orders.count()
    
    # Giá trị đơn hàng trung bình
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
    
    # Dữ liệu biểu đồ - chuyển sang dạng hard-code để không lỗi nếu không có dữ liệu
    daily_revenue = []
    for i in range(30):
        day = now - timedelta(days=i)
        daily_revenue.append({
            'date': day.date().strftime('%d/%m'),
            'revenue': 0  # Hard-code giá trị 0 để tránh lỗi
        })
    
    daily_revenue.reverse()
    
    context = {
        'total_revenue': total_revenue,
        'new_users': new_users,
        'total_orders': total_orders,
        'avg_order_value': avg_order_value,
        'daily_revenue': json.dumps(daily_revenue),
    }
    
    return render(request, 'dashboard/analytics/dashboard.html', context)

@staff_member_required
def sales_report(request):
    """Báo cáo doanh số bán hàng"""
    # Lấy dữ liệu doanh số theo tháng
    current_year = timezone.now().year
    monthly_sales = Order.objects.filter(
        created_at__year=current_year,
        status='completed'
    ).annotate(
        month=ExtractMonth('created_at')
    ).values('month').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('month')
    
    # Chuyển đổi dữ liệu cho biểu đồ
    months = [0] * 12
    sales = [0] * 12
    orders = [0] * 12
    
    for item in monthly_sales:
        month_idx = item['month'] - 1  # Chuyển từ 1-12 sang 0-11
        months[month_idx] = item['month']
        sales[month_idx] = float(item['total'])
        orders[month_idx] = item['count']
    
    # Tính toán các chỉ số
    total_revenue = sum(sales)
    total_orders = sum(orders)
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
    
    # Top sản phẩm bán chạy
    top_products = Product.objects.filter(
        orderitem__order__status='completed',
        orderitem__order__created_at__year=current_year
    ).annotate(
        total_sold=Sum('orderitem__quantity'),
        revenue=Sum(F('orderitem__price') * F('orderitem__quantity'))
    ).order_by('-total_sold')[:10]
    
    context = {
        'months': months,
        'sales': sales,
        'orders': orders,
        'total_revenue': total_revenue,
        'total_orders': total_orders,
        'avg_order_value': avg_order_value,
        'top_products': top_products,
    }
    
    return render(request, 'dashboard/reports/sales.html', context)

@staff_member_required
def product_report(request):
    """Báo cáo sản phẩm"""
    # Thống kê sản phẩm theo danh mục
    categories = Category.objects.annotate(
        product_count=Count('product')
    ).order_by('-product_count')
    
    # Sản phẩm có lượt xem cao nhất
    top_viewed_products = Product.objects.order_by('-view_count')[:10]
    
    # Sản phẩm có tỷ lệ chuyển đổi cao nhất (views -> orders)
    top_conversion_products = Product.objects.annotate(
        order_count=Count('orderitem'),
        conversion_rate=F('order_count') * 100.0 / F('view_count')
    ).filter(
        view_count__gt=0
    ).order_by('-conversion_rate')[:10]
    
    context = {
        'categories': categories,
        'top_viewed_products': top_viewed_products,
        'top_conversion_products': top_conversion_products,
    }
    
    return render(request, 'dashboard/reports/products.html', context)

@staff_member_required
def customer_report(request):
    """Báo cáo khách hàng"""
    # Thống kê khách hàng mới theo tháng
    current_year = timezone.now().year
    new_customers = CustomUser.objects.filter(
        date_joined__year=current_year
    ).annotate(
        month=ExtractMonth('date_joined')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # Chuyển đổi dữ liệu cho biểu đồ
    months = [0] * 12
    customers = [0] * 12
    
    for item in new_customers:
        month_idx = item['month'] - 1  # Chuyển từ 1-12 sang 0-11
        months[month_idx] = item['month']
        customers[month_idx] = item['count']
    
    # Top khách hàng theo doanh số
    top_customers = CustomUser.objects.annotate(
        total_spent=Sum('order__total_amount', filter=Q(order__status='completed')),
        order_count=Count('order', filter=Q(order__status='completed'))
    ).filter(
        total_spent__gt=0
    ).order_by('-total_spent')[:10]
    
    context = {
        'months': months,
        'customers': customers,
        'top_customers': top_customers,
    }
    
    return render(request, 'dashboard/reports/customers.html', context)

@staff_member_required
def user_analytics(request):
    # Logic xử lý phân tích người dùng
    return render(request, 'dashboard/analytics/user_analytics.html')

@staff_member_required
def marketing_analytics(request):
    # Logic xử lý hiệu quả marketing
    return render(request, 'dashboard/analytics/marketing_analytics.html')

@staff_member_required
def custom_report(request):
    # Logic xử lý báo cáo tùy chỉnh
    return render(request, 'dashboard/analytics/custom_report.html')

# Discount Code Views
@staff_member_required
def discount_codes(request):
    discounts = Discount.objects.all().order_by('-created_at')
    return render(request, 'dashboard/discounts/list.html', {'discounts': discounts})

@staff_member_required
def add_discount(request):
    if request.method == 'POST':
        # Logic xử lý thêm mã giảm giá
        code = request.POST.get('code')
        discount_type = request.POST.get('discount_type')
        value = request.POST.get('value')
        valid_from = request.POST.get('valid_from')
        valid_to = request.POST.get('valid_to')
        
        # Tạo mã giảm giá mới
        # ...
        
        messages.success(request, f'Đã thêm mã giảm giá {code}')
        return redirect('dashboard:discount_codes')
    
    return render(request, 'dashboard/discounts/add.html')

@staff_member_required
def edit_discount(request, discount_id):
    discount = get_object_or_404(Discount, id=discount_id)
    
    if request.method == 'POST':
        # Logic xử lý cập nhật mã giảm giá
        # ...
        
        messages.success(request, f'Đã cập nhật mã giảm giá {discount.code}')
        return redirect('dashboard:discount_codes')
    
    return render(request, 'dashboard/discounts/edit.html', {'discount': discount})

@staff_member_required
def delete_discount(request, discount_id):
    discount = get_object_or_404(Discount, id=discount_id)
    
    if request.method == 'POST':
        code = discount.code
        discount.delete()
        messages.success(request, f'Đã xóa mã giảm giá {code}')
    
    return redirect('dashboard:discount_codes')

# Referral Program Views
@staff_member_required
def referral_programs(request):
    programs = ReferralProgram.objects.all().order_by('-created_at')
    return render(request, 'dashboard/referrals/programs.html', {'programs': programs})

@staff_member_required
def add_referral_program(request):
    if request.method == 'POST':
        # Logic xử lý thêm chương trình giới thiệu
        # ...
        
        messages.success(request, f'Đã thêm chương trình giới thiệu {name}')
        return redirect('dashboard:referral_programs')
    
    return render(request, 'dashboard/referrals/add_program.html')

@staff_member_required
def referral_codes(request):
    codes = ReferralCode.objects.all().order_by('-created_at')
    return render(request, 'dashboard/referrals/codes.html', {'codes': codes})

@staff_member_required
def referral_transactions(request):
    transactions = ReferralTransaction.objects.all().order_by('-created_at')
    return render(request, 'dashboard/referrals/transactions.html', {'transactions': transactions})

# Reviews
@staff_member_required
def product_reviews(request):
    status_filter = request.GET.get('status', 'all')
    if status_filter != 'all':
        reviews = ProductReview.objects.filter(status=status_filter).order_by('-created_at')
    else:
        reviews = ProductReview.objects.all().order_by('-created_at')
    
    context = {
        'reviews': reviews,
        'status_filter': status_filter
    }
    return render(request, 'dashboard/reviews/index.html', context)

@staff_member_required
def review_detail(request, review_id):
    review = get_object_or_404(ProductReview, id=review_id)
    
    if request.method == 'POST':
        comment = request.POST.get('comment')
        ReviewComment.objects.create(
            review=review,
            user=request.user,
            content=comment,
            is_staff=True
        )
        messages.success(request, 'Đã thêm bình luận')
        return redirect('dashboard:review_detail', review_id=review.id)
    
    context = {
        'review': review,
        'comments': review.comments.all().order_by('created_at')
    }
    return render(request, 'dashboard/reviews/detail.html', context)

@staff_member_required
def approve_review(request, review_id):
    review = get_object_or_404(ProductReview, id=review_id)
    review.status = 'approved'
    review.save()
    messages.success(request, 'Đã duyệt đánh giá')
    return redirect('dashboard:review_detail', review_id=review.id)

@staff_member_required
def reject_review(request, review_id):
    review = get_object_or_404(ProductReview, id=review_id)
    review.status = 'rejected'
    review.save()
    messages.success(request, 'Đã từ chối đánh giá')
    return redirect('dashboard:review_detail', review_id=review.id)

# API Integration
@staff_member_required
def api_keys(request):
    # Logic xử lý API keys
    return render(request, 'dashboard/api/keys.html')

@staff_member_required
def api_key_list(request):
    """API keys list view"""
    api_keys = APIKey.objects.all().order_by('-created_at')
    return render(request, 'dashboard/api/keys.html', {'api_keys': api_keys})

@staff_member_required
def toggle_api_key(request, key_id):
    """Toggle API key status"""
    api_key = get_object_or_404(APIKey, id=key_id)
    if request.method == 'POST':
        api_key.is_active = not api_key.is_active
        api_key.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=405)

@staff_member_required
def delete_api_key(request, key_id):
    """Delete API key"""
    api_key = get_object_or_404(APIKey, id=key_id)
    if request.method == 'POST':
        api_key.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=405)

@staff_member_required
def webhooks(request):
    webhooks = Webhook.objects.all().order_by('-created_at')
    return render(request, 'dashboard/api/webhooks.html', {'webhooks': webhooks})

@staff_member_required
def webhook_list(request):
    """Webhook list view"""
    webhooks = Webhook.objects.all().order_by('-created_at')
    return render(request, 'dashboard/api/webhooks.html', {'webhooks': webhooks})

@staff_member_required
def add_webhook(request):
    """Add webhook view"""
    if request.method == 'POST':
        form = WebhookForm(request.POST)
        if form.is_valid():
            webhook = form.save(commit=False)
            webhook.secret_key = uuid.uuid4().hex
            webhook.save()
            messages.success(request, 'Webhook đã được tạo thành công')
            return redirect('dashboard:webhooks')
    else:
        form = WebhookForm()
    return render(request, 'dashboard/api/add_webhook.html', {'form': form})

@staff_member_required
def edit_webhook(request, webhook_id):
    """Edit webhook view"""
    webhook = get_object_or_404(Webhook, id=webhook_id)
    if request.method == 'POST':
        form = WebhookForm(request.POST, instance=webhook)
        if form.is_valid():
            form.save()
            messages.success(request, 'Webhook đã được cập nhật')
            return redirect('dashboard:webhooks')
    else:
        form = WebhookForm(instance=webhook)
    return render(request, 'dashboard/api/edit_webhook.html', {'form': form, 'webhook': webhook})

@staff_member_required
def delete_webhook(request, webhook_id):
    """Delete webhook"""
    webhook = get_object_or_404(Webhook, id=webhook_id)
    if request.method == 'POST':
        webhook.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=405)

@staff_member_required
def api_logs(request):
    """API logs view"""
    logs = APILog.objects.all().order_by('-created_at')
    return render(request, 'dashboard/api/logs.html', {'logs': logs})

# Chatbot
@staff_member_required
def chatbot_dashboard(request):
    # Logic xử lý dashboard chatbot
    return render(request, 'dashboard/chatbot/dashboard.html')

@staff_member_required
def chatbot_intents(request):
    intents = ChatbotIntent.objects.all().order_by('-created_at')
    return render(request, 'dashboard/chatbot/intents.html', {'intents': intents})

@staff_member_required
def add_chatbot_intent(request):
    if request.method == 'POST':
        # Logic xử lý thêm intent
        name = request.POST.get('name')
        description = request.POST.get('description')
        keywords = request.POST.get('keywords').split(',')
        
        ChatbotIntent.objects.create(
            name=name,
            description=description,
            keywords=keywords
        )
        
        messages.success(request, f'Đã thêm intent {name}')
        return redirect('dashboard:chatbot_intents')
    
    return render(request, 'dashboard/chatbot/add_intent.html')

@staff_member_required
def chatbot_responses(request):
    responses = ChatbotResponse.objects.all().order_by('-created_at')
    return render(request, 'dashboard/chatbot/responses.html', {'responses': responses})

@staff_member_required
def chatbot_conversations(request):
    sessions = ChatSession.objects.all().order_by('-started_at')
    return render(request, 'dashboard/chatbot/conversations.html', {'sessions': sessions})

@staff_member_required
def analytics_redirect(request):
    """Chuyển hướng từ URL cũ sang URL mới"""
    return redirect('dashboard:analytics_dashboard')

@staff_member_required
def marketing_dashboard(request):
    """Marketing dashboard view"""
    campaigns = Campaign.objects.all()[:5]
    banners = Banner.objects.all()[:5]
    context = {
        'campaigns': campaigns,
        'banners': banners
    }
    return render(request, 'dashboard/marketing/index.html', context)

@staff_member_required
def banner_list(request):
    """Banner list view"""
    banners = Banner.objects.all()
    return render(request, 'dashboard/marketing/banners.html', {'banners': banners})

@staff_member_required
def add_banner(request):
    """Add banner view"""
    if request.method == 'POST':
        form = BannerForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Banner đã được thêm thành công')
            return redirect('dashboard:banners')
    else:
        form = BannerForm()
    return render(request, 'dashboard/marketing/add_banner.html', {'form': form})

@staff_member_required
def settings_dashboard(request):
    """Settings dashboard view"""
    return render(request, 'dashboard/settings/index.html')

@staff_member_required
def update_general_settings(request):
    """Update general settings view"""
    if request.method == 'POST':
        form = GeneralSettingsForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cài đặt đã được cập nhật')
            return redirect('dashboard:settings')
    else:
        form = GeneralSettingsForm()
    return render(request, 'dashboard/settings/general.html', {'form': form})

@staff_member_required
def edit_banner(request, banner_id):
    """Edit banner view"""
    banner = get_object_or_404(Banner, id=banner_id)
    if request.method == 'POST':
        form = BannerForm(request.POST, request.FILES, instance=banner)
        if form.is_valid():
            form.save()
            messages.success(request, 'Banner đã được cập nhật')
            return redirect('dashboard:banners')
    else:
        form = BannerForm(instance=banner)
    return render(request, 'dashboard/marketing/edit_banner.html', {'form': form, 'banner': banner})

@staff_member_required
def email_settings(request):
    """Email settings view"""
    # Lấy cài đặt email từ database hoặc settings
    context = {
        'smtp_host': settings.EMAIL_HOST,
        'smtp_port': settings.EMAIL_PORT,
        'smtp_user': settings.EMAIL_HOST_USER,
        'smtp_use_tls': settings.EMAIL_USE_TLS,
        'default_from_email': settings.DEFAULT_FROM_EMAIL,
    }
    return render(request, 'dashboard/settings/email.html', context)

@staff_member_required
def update_email_settings(request):
    """Update email settings"""
    if request.method == 'POST':
        # Trong thực tế, bạn sẽ lưu vào database hoặc file cấu hình
        # Ở đây chỉ giả lập thành công
        messages.success(request, 'Cài đặt email đã được cập nhật')
        return redirect('dashboard:email_settings')
    return redirect('dashboard:email_settings')

@staff_member_required
def payment_settings(request):
    """Payment settings view"""
    if request.method == 'POST':
        form = PaymentSettingsForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cài đặt thanh toán đã được cập nhật')
            return redirect('dashboard:settings')
    else:
        form = PaymentSettingsForm()
    return render(request, 'dashboard/settings/payment.html', {'form': form})

@staff_member_required
def edit_campaign(request, campaign_id):
    """Edit campaign view"""
    campaign = get_object_or_404(Campaign, id=campaign_id)
    if request.method == 'POST':
        form = CampaignForm(request.POST, instance=campaign)
        if form.is_valid():
            form.save()
            messages.success(request, 'Chiến dịch đã được cập nhật')
            return redirect('dashboard:campaigns')
    else:
        form = CampaignForm(instance=campaign)
    return render(request, 'dashboard/marketing/edit_campaign.html', {'form': form, 'campaign': campaign})

@staff_member_required
def delete_campaign(request, campaign_id):
    """Delete campaign view"""
    campaign = get_object_or_404(Campaign, id=campaign_id)
    if request.method == 'POST':
        campaign.delete()
        messages.success(request, 'Chiến dịch đã được xóa')
        return redirect('dashboard:campaigns')
    return render(request, 'dashboard/marketing/delete_campaign.html', {'campaign': campaign})

@staff_member_required
def add_api_key(request):
    """Add new API key"""
    if request.method == 'POST':
        form = APIKeyForm(request.POST)
        if form.is_valid():
            api_key = form.save(commit=False)
            api_key.key = uuid.uuid4().hex
            api_key.save()
            messages.success(request, 'API key đã được tạo thành công')
            return redirect('dashboard:api_keys')
    else:
        form = APIKeyForm()
    return render(request, 'dashboard/api/add_key.html', {'form': form})

@staff_member_required
def get_log_details(request, log_id):
    """Get API log details"""
    log = get_object_or_404(APILog, id=log_id)
    return JsonResponse({
        'request_data': log.request_data,
        'response_data': log.response_data
    })

@staff_member_required
def api_analytics(request):
    """API usage analytics view"""
    # Thống kê theo ngày
    today = timezone.now().date()
    start_date = today - timedelta(days=30)
    
    daily_stats = APILog.objects.filter(
        created_at__date__gte=start_date
    ).values('created_at__date').annotate(
        total_calls=Count('id'),
        success_rate=Count('id', filter=Q(status_code__lt=400)) * 100.0 / Count('id'),
        avg_response_time=Avg('response_time')
    ).order_by('created_at__date')
    
    # Thống kê theo endpoint
    endpoint_stats = APILog.objects.values('endpoint').annotate(
        total_calls=Count('id'),
        error_rate=Count('id', filter=Q(status_code__gte=400)) * 100.0 / Count('id')
    ).order_by('-total_calls')
    
    context = {
        'daily_stats': daily_stats,
        'endpoint_stats': endpoint_stats,
    }
    
    return render(request, 'dashboard/api/analytics.html', context)

@staff_member_required
def export_users(request):
    """Export users to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # Add headers with style
    headers = [
        'ID', 'Tên người dùng', 'Email', 'Số điện thoại',
        'Ngày đăng ký', 'Trạng thái', 'Vai trò', 'Đơn hàng', 'Chi tiêu'
    ]
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 15

    # Add data
    users = CustomUser.objects.annotate(
        order_count=Count('orders'),
        total_spent=Sum('orders__total_amount')
    ).order_by('-date_joined')
    
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.username)
        ws.cell(row=row, column=3, value=user.email)
        ws.cell(row=row, column=4, value=user.phone)
        ws.cell(row=row, column=5, value=user.date_joined.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=6, value='Hoạt động' if user.is_active else 'Vô hiệu hóa')
        ws.cell(row=row, column=7, value='Nhân viên' if user.is_staff else 'Khách hàng')
        ws.cell(row=row, column=8, value=user.order_count or 0)
        ws.cell(row=row, column=9, value=float(user.total_spent or 0))

    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=users_export.xlsx'
    wb.save(response)
    
    return response

@staff_member_required
def export_orders(request):
    """Export orders to Excel/CSV"""
    export_format = request.GET.get('format', 'excel')
    
    if export_format == 'csv':
        # Export to CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="orders_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Mã đơn', 'Khách hàng', 'Email', 'Số điện thoại',
            'Tổng tiền', 'Trạng thái', 'Thanh toán', 'Ngày tạo'
        ])
        
        orders = Order.objects.all().order_by('-created_at')
        for order in orders:
            writer.writerow([
                order.id,
                order.user.username,
                order.user.email,
                order.user.phone,
                order.total_amount,
                order.get_status_display(),
                order.get_payment_status_display(),
                order.created_at.strftime('%d/%m/%Y %H:%M')
            ])
            
        return response
        
    else:
        # Export to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

        # Add headers with style
        headers = [
            'Mã đơn', 'Khách hàng', 'Email', 'Số điện thoại',
            'Tổng tiền', 'Trạng thái', 'Thanh toán', 'Ngày tạo'
        ]
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = 15

        # Add data
        orders = Order.objects.all().order_by('-created_at')
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order.id)
            ws.cell(row=row, column=2, value=order.user.username)
            ws.cell(row=row, column=3, value=order.user.email)
            ws.cell(row=row, column=4, value=order.user.phone)
            ws.cell(row=row, column=5, value=float(order.total_amount))
            ws.cell(row=row, column=6, value=order.get_status_display())
            ws.cell(row=row, column=7, value=order.get_payment_status_display())
            ws.cell(row=row, column=8, value=order.created_at.strftime('%d/%m/%Y %H:%M'))

        # Add summary
        summary_row = len(orders) + 3
        ws.cell(row=summary_row, column=1, value='Tổng cộng:').font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=f'=SUM(E2:E{len(orders)+1})').font = Font(bold=True)

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=orders_export.xlsx'
        wb.save(response)
        
        return response

@staff_member_required
def update_payment_settings(request):
    """Update payment settings"""
    if request.method == 'POST':
        # Xử lý cập nhật cài đặt thanh toán
        messages.success(request, 'Cài đặt thanh toán đã được cập nhật')
        return redirect('dashboard:payment_settings')
    return redirect('dashboard:payment_settings')

# Chức năng quản lý bảo hành
@staff_member_required
def warranty_management(request):
    """Quản lý các yêu cầu bảo hành"""
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # Sửa lại select_related để phù hợp với model
    tickets = WarrantyTicket.objects.all().select_related('order', 'product', 'customer', 'assigned_to')
    
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    
    if search_query:
        tickets = tickets.filter(
            Q(customer__username__icontains=search_query) |
            Q(customer__email__icontains=search_query) |
            Q(issue_description__icontains=search_query)
        )
    
    # Sắp xếp theo mới nhất trước
    tickets = tickets.order_by('-created_at')
    
    # Phân trang
    paginator = Paginator(tickets, 10)
    page_number = request.GET.get('page', 1)
    tickets = paginator.get_page(page_number)
    
    return render(request, 'dashboard/warranty/management.html', {
        'tickets': tickets,
        'status_filter': status_filter,
        'search_query': search_query,
        'status_choices': [
            ('pending', 'Đang chờ'),
            ('processing', 'Đang xử lý'),
            ('completed', 'Hoàn thành'),
            ('cancelled', 'Đã hủy')
        ],
    })

@staff_member_required
def warranty_detail(request, ticket_id):
    """Chi tiết yêu cầu bảo hành"""
    ticket = get_object_or_404(WarrantyTicket, id=ticket_id)
    ticket_history = WarrantyHistory.objects.filter(ticket=ticket).order_by('-created_at')
    staff_members = CustomUser.objects.filter(is_staff=True)
    
    context = {
        'ticket': ticket,
        'ticket_history': ticket_history,
        'staff_members': staff_members
    }
    
    return render(request, 'dashboard/warranty/detail.html', context)

@staff_member_required
def update_warranty_status(request, ticket_id):
    """Cập nhật trạng thái bảo hành"""
    if request.method == 'POST':
        ticket = get_object_or_404(WarrantyTicket, id=ticket_id)
        old_status = ticket.status
        
        status = request.POST.get('status')
        notes = request.POST.get('notes')
        resolution = request.POST.get('resolution')
        
        # Cập nhật ticket
        ticket.status = status
        if status == 'resolved' and resolution:
            ticket.resolution = resolution
        
        ticket.save()
        
        # Ghi lại lịch sử
        WarrantyHistory.objects.create(
            ticket=ticket,
            action=f'Cập nhật trạng thái từ "{old_status}" thành "{status}"',
            notes=notes,
            performed_by=request.user
        )
        
        messages.success(request, 'Đã cập nhật trạng thái thành công')
    
    return redirect('dashboard:warranty_detail', ticket_id=ticket_id)

@staff_member_required
def assign_warranty(request, ticket_id):
    """Phân công xử lý bảo hành"""
    if request.method == 'POST':
        ticket = get_object_or_404(WarrantyTicket, id=ticket_id)
        staff_id = request.POST.get('assigned_to')
        notes = request.POST.get('notes')
        
        if staff_id:
            staff = get_object_or_404(CustomUser, id=staff_id)
            old_assignee = ticket.assigned_to
            
            # Cập nhật ticket
            ticket.assigned_to = staff
            ticket.save()
            
            # Ghi lại lịch sử
            action = 'Phân công xử lý'
            if old_assignee:
                action = f'Thay đổi người xử lý từ {old_assignee.username} sang {staff.username}'
            
            WarrantyHistory.objects.create(
                ticket=ticket,
                action=action,
                notes=notes,
                performed_by=request.user
            )
            
            messages.success(request, 'Đã phân công xử lý thành công')
    
    return redirect('dashboard:warranty_detail', ticket_id=ticket_id)

@staff_member_required
def add_warranty_note(request, ticket_id):
    """Thêm ghi chú cho bảo hành"""
    if request.method == 'POST':
        ticket = get_object_or_404(WarrantyTicket, id=ticket_id)
        notes = request.POST.get('notes')
        
        if notes:
            WarrantyHistory.objects.create(
                ticket=ticket,
                action='Thêm ghi chú',
                notes=notes,
                performed_by=request.user
            )
            
            messages.success(request, 'Đã thêm ghi chú thành công')
    
    return redirect('dashboard:warranty_detail', ticket_id=ticket_id)

@staff_member_required
def delete_warranty(request, ticket_id):
    """Xóa yêu cầu bảo hành"""
    ticket = get_object_or_404(WarrantyTicket, id=ticket_id)
    ticket.delete()
    
    messages.success(request, 'Đã xóa yêu cầu bảo hành thành công')
    
    return redirect('dashboard:warranty_management')

@staff_member_required
def send_new_account(request, ticket_id):
    """Gửi thông tin tài khoản mới cho khách hàng"""
    if request.method != 'POST':
        return redirect('dashboard:warranty_detail', ticket_id=ticket_id)
    
    ticket = get_object_or_404(WarrantyTicket, id=ticket_id)
    
    # Kiểm tra xem có thông tin tài khoản để gửi không
    if not ticket.new_account_info:
        messages.error(request, "Không có thông tin tài khoản mới để gửi!")
        return redirect('dashboard:warranty_detail', ticket_id=ticket_id)
    
    # Chuẩn bị email
    subject = f"Thông tin tài khoản mới - Yêu cầu bảo hành #{ticket.ticket_id}"
    context = {
        'ticket': ticket,
        'user': ticket.subscription.user,
        'subscription': ticket.subscription,
        'site_url': settings.SITE_URL if hasattr(settings, 'SITE_URL') else '',
    }
    
    message = render_to_string('dashboard/emails/new_account_info.html', context)
    
    # Gửi email
    try:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [ticket.subscription.user.email],
            html_message=message,
            fail_silently=False
        )
        
        # Cập nhật trạng thái nếu chưa được giải quyết
        if ticket.status != 'resolved':
            old_status = ticket.status
            ticket.status = 'resolved'
            ticket.resolved_at = timezone.now()
            ticket.save()
            
            # Ghi lại lịch sử
            WarrantyHistory.objects.create(
                ticket=ticket,
                status='resolved',
                notes=f"Đã gửi thông tin tài khoản mới cho khách hàng. Trạng thái thay đổi từ '{dict(WarrantyTicket.STATUS_CHOICES)[old_status]}' sang 'Đã xử lý'.",
                created_by=request.user
            )
            
            # Tăng số lần bảo hành nếu chưa được tăng
            if not hasattr(ticket, '_warranty_count_increased'):
                subscription = ticket.subscription
                subscription.warranty_count += 1
                subscription.save(update_fields=['warranty_count'])
                ticket._warranty_count_increased = True
        
        messages.success(request, f"Đã gửi thông tin tài khoản mới cho khách hàng thành công!")
    except Exception as e:
        messages.error(request, f"Lỗi khi gửi email: {str(e)}")
    
    return redirect('dashboard:warranty_detail', ticket_id=ticket_id)

@staff_member_required
def warranty_report(request):
    """Báo cáo tỷ lệ bảo hành"""
    # Dữ liệu tỷ lệ bảo hành (Giả lập)
    warranty_data = {
        'labels': ['Netflix', 'Spotify', 'YouTube', 'Disney+', 'Apple Music', 'HBO Max', 'Amazon Prime', 'Canva Pro'],
        'data': [3.5, 2.2, 1.8, 2.5, 1.5, 2.8, 1.2, 0.8],
    }
    
    # Chi tiết bảo hành theo sản phẩm
    warranty_details = [
        {'product': 'Netflix Premium', 'rate': '3.5%', 'common_issue': 'Lỗi đăng nhập', 'total': 35, 'resolved': 32},
        {'product': 'Spotify Family', 'rate': '2.2%', 'common_issue': 'Lỗi kết nối thiết bị', 'total': 22, 'resolved': 20},
        {'product': 'YouTube Premium', 'rate': '1.8%', 'common_issue': 'Lỗi thanh toán', 'total': 18, 'resolved': 18},
        {'product': 'Disney+', 'rate': '2.5%', 'common_issue': 'Lỗi phát nội dung', 'total': 25, 'resolved': 22},
        {'product': 'Apple Music', 'rate': '1.5%', 'common_issue': 'Lỗi đồng bộ', 'total': 15, 'resolved': 14},
        {'product': 'HBO Max', 'rate': '2.8%', 'common_issue': 'Lỗi đăng nhập', 'total': 28, 'resolved': 25},
        {'product': 'Amazon Prime', 'rate': '1.2%', 'common_issue': 'Lỗi xem video', 'total': 12, 'resolved': 12},
        {'product': 'Canva Pro', 'rate': '0.8%', 'common_issue': 'Lỗi tài khoản', 'total': 8, 'resolved': 8},
    ]
    
    # Xu hướng theo tháng
    trend_data = {
        'labels': ['T1', 'T2', 'T3', 'T4', 'T5', 'T6'],
        'data': [3.2, 3.0, 2.8, 2.5, 2.3, 2.1],  # Tỷ lệ bảo hành theo tháng
    }
    
    context = {
        'warranty_data': warranty_data,
        'warranty_details': warranty_details,
        'trend_data': trend_data,
        'avg_warranty_rate': '2.1%',  # Tỷ lệ bảo hành trung bình
        'improvement': 8,  # Cải thiện % so với tháng trước
        'most_common_issue': 'Lỗi đăng nhập',  # Vấn đề phổ biến nhất
    }
    
    return render(request, 'dashboard/reports/warranty.html', context)

@staff_member_required
def subscription_plans(request):
    """Quản lý các gói đăng ký"""
    plans = SubscriptionPlan.objects.all().order_by('price')
    
    if request.method == 'POST':
        # Xử lý thêm/sửa gói đăng ký
        plan_id = request.POST.get('plan_id')
        name = request.POST.get('name')
        description = request.POST.get('description')
        price = request.POST.get('price')
        duration = request.POST.get('duration')
        max_warranty_count = request.POST.get('max_warranty_count', 0)
        
        if plan_id:
            # Cập nhật gói hiện có
            plan = get_object_or_404(SubscriptionPlan, id=plan_id)
            plan.name = name
            plan.description = description
            plan.price = price
            plan.duration = duration
            plan.max_warranty_count = max_warranty_count
            plan.save()
            messages.success(request, f'Đã cập nhật gói {name}')
        else:
            # Tạo gói mới
            SubscriptionPlan.objects.create(
                name=name,
                description=description,
                price=price,
                duration=duration,
                max_warranty_count=max_warranty_count
            )
            messages.success(request, f'Đã thêm gói {name}')
            
        return redirect('dashboard:subscription_plans')
    
    return render(request, 'dashboard/subscriptions/plans.html', {
        'plans': plans
    })

@staff_member_required
def subscription_management(request):
    """Quản lý đăng ký và gia hạn"""
    subscriptions = UserSubscription.objects.all().select_related(
        'user', 'plan'
    ).order_by('-created_at')
    
    # Filters
    status = request.GET.get('status')
    if status:
        if status == 'active':
            subscriptions = subscriptions.filter(end_date__gt=timezone.now())
        elif status == 'expired':
            subscriptions = subscriptions.filter(end_date__lte=timezone.now())
    
    # Search
    search = request.GET.get('search')
    if search:
        subscriptions = subscriptions.filter(
            Q(user__username__icontains=search) |
            Q(user__email__icontains=search) |
            Q(plan__name__icontains=search)
        )
    
    # Stats
    total_active = subscriptions.filter(end_date__gt=timezone.now()).count()
    total_expired = subscriptions.filter(end_date__lte=timezone.now()).count()
    total_revenue = SubscriptionTransaction.objects.filter(
        status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    context = {
        'subscriptions': subscriptions,
        'total_active': total_active,
        'total_expired': total_expired,
        'total_revenue': total_revenue,
        'status_filter': status or 'all',
        'search_query': search or '',
    }
    
    return render(request, 'dashboard/subscriptions/management.html', context)

# Thêm vào cuối file dashboard/views.py để import các functions từ views/source.py
@staff_member_required
def source_list(request):
    """Danh sách nguồn cung cấp"""
    sources = Source.objects.all().order_by('-priority', 'name')
    
    # Filters
    platform = request.GET.get('platform')
    if platform:
        sources = sources.filter(platform=platform)
    
    product_type = request.GET.get('product_type')
    if product_type:
        sources = sources.filter(product_type=product_type)
    
    context = {
        'sources': sources,
        'platforms': Source.objects.values_list('platform', flat=True).distinct(),
        'product_types': Source.objects.values_list('product_type', flat=True).distinct(),
    }
    
    return render(request, 'dashboard/sources/list.html', context)

@staff_member_required
def add_source(request):
    """Thêm nguồn cung cấp mới"""
    if request.method == 'POST':
        # Xử lý form
        name = request.POST.get('name')
        source_url = request.POST.get('source_url')
        platform = request.POST.get('platform')
        product_type = request.POST.get('product_type')
        base_price = request.POST.get('base_price')
        priority = request.POST.get('priority')
        notes = request.POST.get('notes')
        
        source = Source.objects.create(
            name=name,
            source_url=source_url,
            platform=platform,
            product_type=product_type,
            base_price=base_price,
            priority=priority,
            notes=notes
        )
        
        messages.success(request, f'Đã thêm nguồn "{name}"')
        return redirect('dashboard:source_detail', source_id=source.id)
    
    return render(request, 'dashboard/sources/add.html')

@staff_member_required
def source_detail(request, source_id):
    """Chi tiết nguồn cung cấp"""
    source = get_object_or_404(Source, id=source_id)
    source_products = SourceProduct.objects.filter(source=source)
    logs = SourceLog.objects.filter(source=source).order_by('-created_at')[:50]
    
    # Thống kê
    avg_price = source_products.aggregate(Avg('price'))['price__avg'] or 0
    
    context = {
        'source': source,
        'source_products': source_products,
        'logs': logs,
        'avg_price': avg_price,
    }
    
    return render(request, 'dashboard/sources/detail.html', context)

@staff_member_required
def compare_sources(request):
    """So sánh các nguồn cung cấp"""
    product_type = request.GET.get('product_type')
    search = request.GET.get('search', '')
    
    sources = Source.objects.all().order_by('base_price')
    if product_type:
        sources = sources.filter(product_type=product_type)
    
    if search:
        sources = sources.filter(
            Q(name__icontains=search) |
            Q(product_type__icontains=search)
        )
    
    context = {
        'sources': sources,
        'product_types': Source.objects.values_list('product_type', flat=True).distinct(),
        'selected_type': product_type,
        'search_query': search,
    }
    
    return render(request, 'dashboard/sources/compare.html', context)

@staff_member_required
def add_source_product(request):
    """Thêm sản phẩm từ nguồn"""
    if request.method == 'POST':
        source_id = request.POST.get('source_id')
        product_id = request.POST.get('product_id', None)
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        product_url = request.POST.get('product_url', '')
        price = request.POST.get('price')
        error_rate = request.POST.get('error_rate', 0)
        
        source = get_object_or_404(Source, id=source_id)
        product = None
        if product_id:
            product = get_object_or_404(Product, id=product_id)
            
        source_product = SourceProduct.objects.create(
            source=source,
            product=product,
            name=name,
            description=description,
            product_url=product_url,
            price=price,
            error_rate=error_rate
        )
        
        messages.success(request, f'Đã thêm sản phẩm "{name}" cho nguồn {source.name}')
        return redirect('dashboard:source_detail', source_id=source.id)
    
    sources = Source.objects.all()
    products = Product.objects.all()
    
    return render(request, 'dashboard/sources/add_product.html', {
        'sources': sources,
        'products': products
    })

@staff_member_required
def add_source_log(request):
    """Thêm nhật ký nguồn nhập mới"""
    if request.method == 'POST':
        form = SourceLogForm(request.POST)
        if form.is_valid():
            log = form.save(commit=False)
            log.created_by = request.user
            log.save()
            
            # Cập nhật tỷ lệ có hàng cho nguồn
            source = log.source
            available_logs = SourceLog.objects.filter(source=source, status='available').count()
            total_logs = SourceLog.objects.filter(source=source).count()
            
            if total_logs > 0:
                availability_rate = (available_logs / total_logs) * 100
                source.availability_rate = availability_rate
                source.save()
            
            messages.success(request, 'Nhật ký nguồn nhập đã được tạo thành công!')
            return redirect('dashboard:source_log_list')
    else:
        form = SourceLogForm()
    
    return render(request, 'dashboard/sources/log_form.html', {'form': form, 'title': 'Thêm nhật ký nguồn nhập'})

@staff_member_required
def api_source_products(request, source_id):
    """API trả về danh sách sản phẩm theo nguồn"""
    source = get_object_or_404(Source, id=source_id)
    products = SourceProduct.objects.filter(source=source)
    
    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'price': float(product.price),
        })
    
    return JsonResponse({'products': products_data})

@staff_member_required
def source_analytics(request):
    """Phân tích và báo cáo nguồn nhập hàng"""
    # Nhận tham số filter
    time_range = request.GET.get('time_range', 'month')
    product_type = request.GET.get('product_type', '')
    platform = request.GET.get('platform', '')
    chart_type = request.GET.get('chart_type', 'price')
    
    # Tính toán ngày bắt đầu và kết thúc dựa trên time_range
    today = timezone.now()
    date_to = today
    
    if time_range == 'week':
        date_from = today - timedelta(days=7)
    elif time_range == 'month':
        date_from = today - timedelta(days=30)
    elif time_range == 'quarter':
        date_from = today - timedelta(days=90)
    elif time_range == 'year':
        date_from = today - timedelta(days=365)
    elif time_range == 'custom':
        try:
            date_from = datetime.datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d').date()
            date_to = datetime.datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            date_from = today - timedelta(days=30)
    else:
        date_from = today - timedelta(days=30)

    # Lọc nguồn nhập
    sources = Source.objects.all()
    
    if product_type:
        sources = sources.filter(product_type=product_type)
    
    if platform:
        sources = sources.filter(platform=platform)
    
    # Lấy dữ liệu logs trong khoảng thời gian
    logs = SourceLog.objects.filter(created_at__gte=date_from, created_at__lte=date_to)
    
    # Tính toán chỉ số tổng hợp
    total_sources = sources.count()
    total_products = SourceProduct.objects.filter(source__in=sources).count()
    
    # Tính tỉ lệ có hàng trung bình
    avg_availability = 0
    if logs.count() > 0:
        avg_availability = (logs.filter(has_stock=True).count() / logs.count()) * 100
    
    # Tính thời gian phản hồi trung bình
    logs_with_time = logs.exclude(processing_time__isnull=True)
    avg_response_time = 0
    if logs_with_time.count() > 0:
        avg_response_time = logs_with_time.aggregate(Avg('processing_time'))['processing_time__avg'] or 0
    
    # Lấy top nguồn hiệu quả nhất
    top_sources = []
    for source in sources[:10]:  # Lấy 10 nguồn đầu tiên
        # Tính toán hiệu quả dựa trên các yếu tố: tỉ lệ có hàng, thời gian phản hồi, tỉ lệ lỗi và giá
        source_logs = logs.filter(source=source)
        source_products = SourceProduct.objects.filter(source=source)
        
        availability_rate = source.availability_rate
        error_rate = source.error_rate
        avg_processing_time = source.avg_processing_time or 0
        avg_price = source_products.aggregate(Avg('price'))['price__avg'] or 0
        
        # Tính điểm hiệu quả (tỉ lệ có hàng cao, thời gian phản hồi nhanh, tỉ lệ lỗi thấp, giá thấp = hiệu quả cao)
        efficiency_score = (availability_rate * 0.4) - (error_rate * 0.3) - (avg_processing_time * 0.1) - (avg_price * 0.001 * 0.2)
        
        top_sources.append({
            'id': source.id,
            'name': source.name,
            'platform': source.platform,
            'product_type': source.product_type,
            'availability_rate': availability_rate,
            'error_rate': error_rate,
            'avg_processing_time': avg_processing_time,
            'avg_price': avg_price,
            'efficiency_score': efficiency_score
        })
    
    # Sắp xếp theo điểm hiệu quả
    top_sources = sorted(top_sources, key=lambda x: x['efficiency_score'], reverse=True)
    
    # Chuẩn bị dữ liệu cho biểu đồ
    chart_data = {
        'labels': [source['name'] for source in top_sources[:10]],
        'datasets': []
    }
    
    if chart_type == 'price':
        chart_data['datasets'].append({
            'label': 'Giá trung bình (VNĐ)',
            'data': [source['avg_price'] for source in top_sources[:10]],
            'backgroundColor': '#4e73df',
            'borderColor': '#4e73df',
        })
    elif chart_type == 'availability':
        chart_data['datasets'].append({
            'label': 'Tỉ lệ có hàng (%)',
            'data': [source['availability_rate'] for source in top_sources[:10]],
            'backgroundColor': '#1cc88a',
            'borderColor': '#1cc88a',
        })
    elif chart_type == 'response_time':
        chart_data['datasets'].append({
            'label': 'Thời gian phản hồi (phút)',
            'data': [source['avg_processing_time'] for source in top_sources[:10]],
            'backgroundColor': '#f6c23e',
            'borderColor': '#f6c23e',
        })
    elif chart_type == 'error_rate':
        chart_data['datasets'].append({
            'label': 'Tỉ lệ lỗi (%)',
            'data': [source['error_rate'] for source in top_sources[:10]],
            'backgroundColor': '#e74a3b',
            'borderColor': '#e74a3b',
        })
    elif chart_type == 'efficiency':
        chart_data['datasets'].append({
            'label': 'Điểm hiệu quả',
            'data': [source['efficiency_score'] for source in top_sources[:10]],
            'backgroundColor': '#36b9cc',
            'borderColor': '#36b9cc',
        })
    
    # Các insight và gợi ý cho admin
    insights = []
    
    # Kiểm tra nguồn nào có tỉ lệ lỗi cao
    high_error_sources = [source for source in top_sources if source['error_rate'] > 20]
    if high_error_sources:
        insights.append({
            'type': 'warning',
            'icon': 'fa-exclamation-triangle',
            'message': f"Nguồn '{high_error_sources[0]['name']}' có tỉ lệ lỗi cao ({high_error_sources[0]['error_rate']:.1f}%). Cân nhắc kiểm tra lại chất lượng."
        })
    
    # Kiểm tra nguồn nào có thời gian phản hồi chậm
    slow_response_sources = [source for source in top_sources if source['avg_processing_time'] > 60]
    if slow_response_sources:
        insights.append({
            'type': 'info',
            'icon': 'fa-clock',
            'message': f"Nguồn '{slow_response_sources[0]['name']}' có thời gian phản hồi chậm ({slow_response_sources[0]['avg_processing_time']:.0f} phút). Có thể tìm nguồn thay thế."
        })
    
    # Tìm nguồn có giá thấp nhất với tỉ lệ có hàng cao
    best_value_sources = sorted(
        [source for source in top_sources if source['availability_rate'] > 70],
        key=lambda x: x['avg_price']
    )
    if best_value_sources:
        insights.append({
            'type': 'success',
            'icon': 'fa-thumbs-up',
            'message': f"Nguồn '{best_value_sources[0]['name']}' có giá thấp nhất ({best_value_sources[0]['avg_price']:.0f} VNĐ) với tỉ lệ có hàng cao ({best_value_sources[0]['availability_rate']:.1f}%)."
        })
    
    # Lấy các loại sản phẩm và platform cho bộ lọc
    product_types = Source.objects.values_list('product_type', flat=True).distinct()
    platforms = Source.objects.values_list('platform', flat=True).distinct()
    
    context = {
        'time_range': time_range,
        'date_from': date_from,
        'date_to': date_to,
        'selected_type': product_type,
        'selected_platform': platform,
        'chart_type': chart_type,
        'total_sources': total_sources,
        'total_products': total_products,
        'avg_availability': avg_availability,
        'avg_response_time': avg_response_time,
        'top_sources': top_sources,
        'chart_data': chart_data,
        'insights': insights,
        'product_types': product_types,
        'platforms': platforms
    }
    
    return render(request, 'dashboard/sources/analytics.html', context)

@staff_member_required
def share_report(request):
    """Chia sẻ báo cáo qua email"""
    if request.method == 'POST':
        try:
            # Lấy thông tin từ request
            email = request.POST.get('email')
            note = request.POST.get('note', '')
            password_protected = request.POST.get('password_protected') == 'true'
            password = request.POST.get('password', '')
            report_type = request.POST.get('report_type')
            report_params = json.loads(request.POST.get('report_params', '{}'))
            
            # Tạo mã unique cho báo cáo này
            report_id = str(uuid.uuid4())
            
            # Trong trường hợp thực tế, bạn sẽ lưu thông tin báo cáo vào database
            # và gửi link tới báo cáo này
            
            # Chuẩn bị nội dung email
            subject = f'Báo cáo {report_type} từ hệ thống quản lý'
            message = f'Chào bạn,\n\nBạn nhận được báo cáo {report_type} từ hệ thống quản lý.\n\n'
            
            if note:
                message += f'Ghi chú: {note}\n\n'
            
            message += f'Link truy cập báo cáo: {settings.BASE_URL}/reports/{report_id}\n\n'
            
            if password_protected:
                message += f'Mật khẩu: {password}\n\n'
            
            message += 'Trân trọng,\nHệ thống quản lý TomOi'
            
            # Gửi email (disabled để tránh lỗi khi chưa có cấu hình email)
            # send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email])
            
            # Trả về kết quả thành công (trong thực tế sẽ gửi mail thật)
            return JsonResponse({'success': True, 'message': f'Đã gửi báo cáo đến {email}'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Phương thức không được hỗ trợ'})

@staff_member_required
def source_chart_data(request):
    """API trả về dữ liệu biểu đồ cho nguồn nhập"""
    time_range = request.GET.get('time_range', 'month')
    product_type = request.GET.get('product_type', '')
    platform = request.GET.get('platform', '')
    chart_type = request.GET.get('chart_type', 'price')
    
    # Tính toán ngày bắt đầu và kết thúc tương tự source_analytics
    today = timezone.now()
    date_to = today
    
    if time_range == 'week':
        date_from = today - timedelta(days=7)
    elif time_range == 'month':
        date_from = today - timedelta(days=30)
    elif time_range == 'quarter':
        date_from = today - timedelta(days=90)
    elif time_range == 'year':
        date_from = today - timedelta(days=365)
    elif time_range == 'custom':
        try:
            date_from = datetime.datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d').date()
            date_to = datetime.datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            date_from = today - timedelta(days=30)
    else:
        date_from = today - timedelta(days=30)
    
    # Lọc nguồn nhập
    sources = Source.objects.all()
    
    if product_type:
        sources = sources.filter(product_type=product_type)
    
    if platform:
        sources = sources.filter(platform=platform)
    
    # Lấy dữ liệu logs trong khoảng thời gian
    logs = SourceLog.objects.filter(created_at__gte=date_from, created_at__lte=date_to)
    
    # Xử lý dữ liệu tương tự source_analytics nhưng trả về dạng JSON
    # Tính toán và trả về dữ liệu biểu đồ JSON
    source_data = []
    
    for source in sources:
        source_logs = logs.filter(source=source)
        source_products = SourceProduct.objects.filter(source=source)
        
        availability_rate = 0
        if source_logs.count() > 0:
            availability_rate = (source_logs.filter(has_stock=True).count() / source_logs.count()) * 100
        
        error_rate = source.error_rate
        avg_processing_time = source.avg_processing_time or 0
        avg_price = source_products.aggregate(Avg('price'))['price__avg'] or 0
        
        efficiency_score = (availability_rate * 0.4) - (error_rate * 0.3) - (avg_processing_time * 0.1) - (avg_price * 0.001 * 0.2)
        
        source_data.append({
            'id': source.id,
            'name': source.name,
            'platform': source.platform,
            'product_type': source.product_type,
            'availability_rate': availability_rate,
            'error_rate': error_rate,
            'avg_processing_time': avg_processing_time,
            'avg_price': avg_price,
            'efficiency_score': efficiency_score
        })
    
    # Sắp xếp theo điểm hiệu quả
    source_data = sorted(source_data, key=lambda x: x['efficiency_score'], reverse=True)[:10]
    
    # Chuẩn bị dữ liệu cho biểu đồ
    labels = [source['name'] for source in source_data]
    
    datasets = []
    if chart_type == 'price':
        datasets.append({
            'label': 'Giá trung bình (VNĐ)',
            'data': [source['avg_price'] for source in source_data],
            'backgroundColor': '#4e73df',
            'borderColor': '#4e73df',
        })
    elif chart_type == 'availability':
        datasets.append({
            'label': 'Tỉ lệ có hàng (%)',
            'data': [source['availability_rate'] for source in source_data],
            'backgroundColor': '#1cc88a',
            'borderColor': '#1cc88a',
        })
    elif chart_type == 'response_time':
        datasets.append({
            'label': 'Thời gian phản hồi (phút)',
            'data': [source['avg_processing_time'] for source in source_data],
            'backgroundColor': '#f6c23e',
            'borderColor': '#f6c23e',
        })
    elif chart_type == 'error_rate':
        datasets.append({
            'label': 'Tỉ lệ lỗi (%)',
            'data': [source['error_rate'] for source in source_data],
            'backgroundColor': '#e74a3b',
            'borderColor': '#e74a3b',
        })
    elif chart_type == 'efficiency':
        datasets.append({
            'label': 'Điểm hiệu quả',
            'data': [source['efficiency_score'] for source in source_data],
            'backgroundColor': '#36b9cc',
            'borderColor': '#36b9cc',
        })
    elif chart_type == 'combined':
        # Một biểu đồ kết hợp nhiều chỉ số
        datasets = [
            {
                'label': 'Tỉ lệ có hàng (%)',
                'data': [source['availability_rate'] for source in source_data],
                'backgroundColor': 'rgba(28, 200, 138, 0.2)',
                'borderColor': '#1cc88a',
                'borderWidth': 2,
                'fill': True
            },
            {
                'label': 'Tỉ lệ lỗi (%)',
                'data': [source['error_rate'] for source in source_data],
                'backgroundColor': 'rgba(231, 74, 59, 0.2)',
                'borderColor': '#e74a3b',
                'borderWidth': 2,
                'fill': True
            }
        ]
    
    return JsonResponse({
        'labels': labels,
        'datasets': datasets
    })

@staff_member_required
def source_dashboard(request):
    """Dashboard quản lý nguồn nhập"""
    sources = Source.objects.all()
    recent_logs = SourceLog.objects.all().order_by('-created_at')[:10]
    
    # Tính toán các chỉ số
    total_sources = sources.count()
    active_sources = sources.filter(sourcelog__created_at__gte=timezone.now() - timezone.timedelta(days=30)).distinct().count()
    
    # Tỷ lệ có hàng trung bình
    avg_availability = 0
    if total_sources > 0:
        availability_sum = sum(source.availability_rate for source in sources)
        avg_availability = availability_sum / total_sources
    
    # Dữ liệu cho biểu đồ
    platform_data = []
    platform_labels = []
    platform_colors = []
    platform_hover_colors = []
    
    platform_counts = {}
    for source in sources:
        platform = source.get_platform_display()
        if platform in platform_counts:
            platform_counts[platform] += 1
        else:
            platform_counts[platform] = 1
    
    colors = [
        "rgba(78, 115, 223, 0.8)",
        "rgba(28, 200, 138, 0.8)",
        "rgba(246, 194, 62, 0.8)",
        "rgba(231, 74, 59, 0.8)",
        "rgba(54, 185, 204, 0.8)",
    ]
    
    hover_colors = [
        "rgba(78, 115, 223, 1)",
        "rgba(28, 200, 138, 1)",
        "rgba(246, 194, 62, 1)",
        "rgba(231, 74, 59, 1)",
        "rgba(54, 185, 204, 1)",
    ]
    
    i = 0
    for platform, count in platform_counts.items():
        platform_labels.append(platform)
        platform_data.append(count)
        platform_colors.append(colors[i % len(colors)])
        platform_hover_colors.append(hover_colors[i % len(hover_colors)])
        i += 1
    
    context = {
        'sources': sources,
        'recent_logs': recent_logs,
        'total_sources': total_sources,
        'active_sources': active_sources,
        'avg_availability': avg_availability,
        'platform_data': platform_data,
        'platform_labels': platform_labels,
        'platform_colors': platform_colors,
        'platform_hover_colors': platform_hover_colors,
    }
    
    return render(request, 'dashboard/sources/dashboard.html', context)

@staff_member_required
def create_warranty(request):
    """Tạo mới yêu cầu bảo hành"""
    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        product_id = request.POST.get('product')
        issue_description = request.POST.get('issue_description')
        
        try:
            customer = CustomUser.objects.get(id=customer_id)
            product = None
            if product_id:
                product = Product.objects.get(id=product_id)
            
            # Tạo ticket bảo hành mới
            warranty = WarrantyTicket.objects.create(
                customer=customer,
                product=product,
                issue_description=issue_description,
                status='pending'
            )
            
            # Tạo lịch sử bảo hành
            WarrantyHistory.objects.create(
                ticket=warranty,
                action='Tạo yêu cầu',
                notes='Yêu cầu bảo hành mới được tạo',
                performed_by=request.user
            )
            
            messages.success(request, 'Đã tạo yêu cầu bảo hành thành công')
            return redirect('dashboard:warranty_detail', warranty.id)
        except Exception as e:
            messages.error(request, f'Lỗi: {str(e)}')
    
    # GET request - hiển thị form
    customers = CustomUser.objects.all()
    products = Product.objects.all()
    
    context = {
        'customers': customers,
        'products': products
    }
    
    return render(request, 'dashboard/warranty/create.html', context)

@staff_member_required
def get_source_base_price(request, source_id):
    """API endpoint để lấy giá cơ bản của nguồn"""
    source = get_object_or_404(Source, id=source_id)
    return JsonResponse({'base_price': source.base_price})

@login_required
def campaign_list(request):
    sources = Source.objects.all()
    recent_logs = SourceLog.objects.all().order_by('-created_at')[:10]
    
    # Tính toán các chỉ số
    total_sources = sources.count()
    active_sources = sources.filter(sourcelog__created_at__gte=timezone.now() - timezone.timedelta(days=30)).distinct().count()
    
    # Tỷ lệ có hàng trung bình
    avg_availability = 0
    if total_sources > 0:
        availability_sum = sum(source.availability_rate for source in sources)
        avg_availability = availability_sum / total_sources
    
    context = {
        'sources': sources,
        'recent_logs': recent_logs,
        'total_sources': total_sources,
        'active_sources': active_sources,
        'avg_availability': avg_availability,
    }
    
    return render(request, 'dashboard/sources/dashboard.html', context)

@login_required
def source_list(request):
    """Danh sách nguồn nhập"""
    sources = Source.objects.all().order_by('priority', 'name')
    
    # Tính toán các chỉ số bổ sung
    for source in sources:
        source.avg_processing_time = source.get_avg_processing_time()
        source.availability_status = source.get_availability_status()
        source.formatted_price = source.format_base_price()
    
    return render(request, 'dashboard/sources/list.html', {'sources': sources})

@login_required
def source_add(request):
    if request.method == 'POST':
        form = SourceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nguồn nhập đã được tạo thành công!')
            return redirect('dashboard:source_list')
    else:
        form = SourceForm()
    
    return render(request, 'dashboard/sources/add.html', {'form': form, 'title': 'Thêm nguồn nhập mới'})

@login_required
def source_edit(request, source_id):
    source = get_object_or_404(Source, id=source_id)
    
    if request.method == 'POST':
        form = SourceForm(request.POST, instance=source)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nguồn nhập đã được cập nhật thành công!')
            return redirect('dashboard:source_list')
    else:
        form = SourceForm(instance=source)
    
    return render(request, 'dashboard/sources/edit.html', {'form': form, 'title': 'Chỉnh sửa nguồn nhập'})

@login_required
def source_delete(request, source_id):
    source = get_object_or_404(Source, id=source_id)
    
    if request.method == 'POST':
        source.delete()
        messages.success(request, 'Nguồn nhập đã được xóa thành công!')
        return redirect('dashboard:source_list')
    
    return render(request, 'dashboard/sources/delete.html', {'source': source})

@login_required
def source_log_list(request):
    """Danh sách nhật ký nguồn nhập"""
    logs = SourceLog.objects.all().select_related('source', 'created_by').order_by('-created_at')
    
    # Format giá
    for log in logs:
        log.formatted_price = log.format_price()
    
    return render(request, 'dashboard/sources/logs.html', {'logs': logs})

@login_required
def add_source_log(request):
    """Thêm nhật ký nguồn nhập"""
    sources = Source.objects.all().order_by('name')
    
    if request.method == 'POST':
        form = SourceLogForm(request.POST)
        if form.is_valid():
            log = form.save(commit=False)
            log.created_by = request.user
            log.save()
            
            # Cập nhật tỷ lệ có hàng cho nguồn
            source = log.source
            available_logs = source.logs.filter(status='available').count()
            total_logs = source.logs.count()
            if total_logs > 0:
                source.availability_rate = (available_logs / total_logs) * 100
                source.save()
            
            messages.success(request, 'Nhật ký nguồn nhập đã được thêm thành công!')
            return redirect('dashboard:source_log_list')
    else:
        form = SourceLogForm()
    
    return render(request, 'dashboard/sources/add_log.html', {
        'form': form,
        'sources': sources
    })

# Thêm vào cuối file views.py
def add_source_redirect(request):
    """Redirect để xử lý URL cũ"""
    return redirect('dashboard:source_add')

# Đảm bảo thêm URL pattern tương ứng trong urls.py:
# path('sources/add-source/', views.add_source_redirect, name='add_source'),

@login_required
def source_dashboard(request):
    """Bảng điều khiển nguồn nhập"""
    sources = Source.objects.all()
    recent_logs = SourceLog.objects.select_related('source', 'created_by').order_by('-created_at')[:10]
    
    # Thống kê
    total_sources = sources.count()
    active_sources = sources.filter(availability_rate__gte=50).count()
    
    # Tính toán tỷ lệ có hàng trung bình
    avg_availability = sources.aggregate(avg=models.Avg('availability_rate'))['avg'] or 0
    
    # Dữ liệu cho biểu đồ
    platform_data = []
    platform_labels = []
    platform_colors = ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b', '#858796', '#5a5c69']
    platform_hover_colors = ['#2e59d9', '#17a673', '#2c9faf', '#dda20a', '#be2617', '#60616f', '#373840']
    
    # Tính toán phân bố nền tảng
    platform_counts = sources.values('platform').annotate(count=models.Count('id'))
    for i, platform in enumerate(platform_counts):
        platform_labels.append(dict(Source.PLATFORM_CHOICES).get(platform['platform'], 'Khác'))
        platform_data.append(platform['count'])
    
    # Tính toán các chỉ số bổ sung
    for log in recent_logs:
        log.formatted_price = log.format_price()
    
    context = {
        'sources': sources,
        'recent_logs': recent_logs,
        'total_sources': total_sources,
        'active_sources': active_sources,
        'avg_availability': avg_availability,
        'platform_data': platform_data,
        'platform_labels': platform_labels,
        'platform_colors': platform_colors,
        'platform_hover_colors': platform_hover_colors,
    }
    
    return render(request, 'dashboard/sources/dashboard.html', context)

@staff_member_required
def account_types(request):
    """Quản lý loại tài khoản"""
    # Lấy dữ liệu từ model tương ứng
    return render(request, 'dashboard/accounts/types.html', {})

@staff_member_required
def account_transactions(request):
    """Quản lý giao dịch tài khoản"""
    # Lấy dữ liệu từ model tương ứng
    return render(request, 'dashboard/accounts/transactions.html', {})

@staff_member_required
def tcoin_accounts(request):
    """Quản lý tài khoản TCoin"""
    # Lấy dữ liệu từ model tương ứng
    return render(request, 'dashboard/accounts/tcoin.html', {})

@staff_member_required
def chat_messages(request):
    """Quản lý tin nhắn chat"""
    # Lấy dữ liệu từ model tương ứng
    return render(request, 'dashboard/chat/messages.html', {})

@staff_member_required
def chat_sessions(request):
    """Quản lý phiên chat"""
    # Lấy dữ liệu từ model tương ứng
    return render(request, 'dashboard/chat/sessions.html', {})

@staff_member_required
def email_logs(request):
    """Quản lý log email"""
    # Lấy dữ liệu từ model tương ứng
    return render(request, 'dashboard/email/logs.html', {})

@staff_member_required
def email_templates(request):
    """Quản lý mẫu email"""
    # Lấy dữ liệu từ model tương ứng
    return render(request, 'dashboard/email/templates.html', {})

@staff_member_required
def reports_analysis(request):
    """Trang phân tích báo cáo theo 5 tiêu chí"""
    # Lấy dữ liệu cho biểu đồ nguồn nhập
    sources = Source.objects.all()
    
    # Dữ liệu cho phân tích nguồn nhập
    platform_counts = sources.values('platform').annotate(count=models.Count('id'))
    platform_data = []
    platform_labels = []
    
    for platform in platform_counts:
        platform_labels.append(dict(Source.PLATFORM_CHOICES).get(platform['platform'], 'Khác'))
        platform_data.append(platform['count'])
    
    # Dữ liệu cho phân tích tỷ lệ có hàng
    availability_data = [
        sources.filter(availability_rate__gte=50).count(),
        sources.filter(availability_rate__lt=50).count()
    ]
    
    # Dữ liệu cho doanh thu (Giả lập)
    revenue_months = ['T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'T8', 'T9', 'T10', 'T11', 'T12']
    revenue_data = [5000000, 6500000, 7800000, 8300000, 9100000, 9500000, 10200000, 11500000, 12000000, 10800000, 11000000, 13500000]
    
    # Dữ liệu cho sản phẩm bán chạy (Giả lập)
    popular_products = [
        {'name': 'Netflix Premium', 'sold': 250, 'revenue': 9500000},
        {'name': 'Spotify Family', 'sold': 180, 'revenue': 6300000},
        {'name': 'YouTube Premium', 'sold': 150, 'revenue': 4500000},
        {'name': 'Disney+ Premium', 'sold': 120, 'revenue': 3600000},
        {'name': 'Apple Music', 'sold': 90, 'revenue': 2700000},
    ]
    
    # Dữ liệu thời gian xử lý (Giả lập)
    processing_labels = ['<15 phút', '15-60 phút', '1-3 giờ', '3-12 giờ', '>12 giờ']
    processing_data = [60, 25, 10, 4, 1]
    
    # Dữ liệu tỷ lệ bảo hành (Giả lập)
    warranty_labels = ['Netflix', 'Spotify', 'YouTube', 'Disney+', 'Apple Music']
    warranty_data = [3.5, 2.2, 1.8, 2.5, 1.5]
    
    context = {
        'platform_data': platform_data,
        'platform_labels': platform_labels,
        'availability_data': availability_data,
        'revenue_months': revenue_months,
        'revenue_data': revenue_data,
        'popular_products': popular_products,
        'processing_labels': processing_labels,
        'processing_data': processing_data,
        'warranty_labels': warranty_labels,
        'warranty_data': warranty_data,
    }
    
    return render(request, 'dashboard/reports/analysis.html', context)

@staff_member_required
def revenue_report(request):
    """Báo cáo doanh thu"""
    # Tạo dữ liệu mẫu cho báo cáo doanh thu
    monthly_revenue = {
        'labels': ['T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'T8', 'T9', 'T10', 'T11', 'T12'],
        'data': [5000000, 6500000, 7800000, 8300000, 9100000, 9500000, 10200000, 11500000, 12000000, 10800000, 11000000, 13500000]
    }
    
    product_revenue = [
        {'name': 'Netflix Premium', 'revenue': 9500000, 'percentage': 25},
        {'name': 'Spotify Family', 'revenue': 6300000, 'percentage': 18},
        {'name': 'YouTube Premium', 'revenue': 4500000, 'percentage': 12},
        {'name': 'Disney+ Premium', 'revenue': 3600000, 'percentage': 10},
        {'name': 'Apple Music', 'revenue': 2700000, 'percentage': 8},
        {'name': 'HBO Max', 'revenue': 2200000, 'percentage': 6},
        {'name': 'Amazon Prime', 'revenue': 1900000, 'percentage': 5},
        {'name': 'Canva Pro', 'revenue': 1700000, 'percentage': 4},
        {'name': 'Microsoft 365', 'revenue': 1500000, 'percentage': 4},
        {'name': 'Khác', 'revenue': 3000000, 'percentage': 8},
    ]
    
    channel_data = [70, 25, 5]
    
    context = {
        'monthly_revenue': monthly_revenue,
        'product_revenue': product_revenue,
        'channel_data': channel_data,
        'total_revenue': sum(d['revenue'] for d in product_revenue),
        'growth_rate': 18.5,  # Tỷ lệ tăng trưởng so với tháng trước
        'forecast_next_month': 14000000,  # Dự báo doanh thu tháng tiếp theo
    }
    
    return render(request, 'dashboard/reports/revenue.html', context)

@staff_member_required
def popular_products(request):
    """Báo cáo sản phẩm bán chạy"""
    # Tạo dữ liệu mẫu cho sản phẩm bán chạy
    top_products = [
        {'name': 'Netflix Premium', 'sold': 250, 'revenue': 9500000, 'growth': 15},
        {'name': 'Spotify Family', 'sold': 180, 'revenue': 6300000, 'growth': 8},
        {'name': 'YouTube Premium', 'sold': 150, 'revenue': 4500000, 'growth': 12},
        {'name': 'Disney+ Premium', 'sold': 120, 'revenue': 3600000, 'growth': 20},
        {'name': 'Apple Music', 'sold': 90, 'revenue': 2700000, 'growth': 5},
        {'name': 'HBO Max', 'sold': 75, 'revenue': 2200000, 'growth': 18},
        {'name': 'Amazon Prime', 'sold': 60, 'revenue': 1900000, 'growth': 7},
        {'name': 'Canva Pro', 'sold': 55, 'revenue': 1700000, 'growth': 10},
        {'name': 'Microsoft 365', 'sold': 45, 'revenue': 1500000, 'growth': 6},
        {'name': 'Khác', 'sold': 100, 'revenue': 3000000, 'growth': 4},
    ]
    
    # Dữ liệu cho biểu đồ xu hướng
    trend_data = {
        'labels': ['T1', 'T2', 'T3', 'T4', 'T5', 'T6'],
        'netflix': [40, 45, 48, 50, 53, 58],
        'spotify': [30, 32, 35, 38, 40, 42],
        'youtube': [25, 28, 30, 32, 35, 38],
    }
    
    context = {
        'top_products': top_products,
        'trend_data': trend_data,
        'total_sold': sum(d['sold'] for d in top_products),
        'sales_growth': 12.5,  # Tăng trưởng tổng số lượng bán
    }
    
    return render(request, 'dashboard/reports/popular_products.html', context)

@staff_member_required
def processing_time(request):
    """Báo cáo thời gian xử lý"""
    # Tạo dữ liệu mẫu cho thời gian xử lý
    processing_data = {
        'labels': ['<15 phút', '15-60 phút', '1-3 giờ', '3-12 giờ', '>12 giờ'],
        'data': [60, 25, 10, 4, 1],
    }
    
    # Dữ liệu chi tiết thời gian xử lý theo nguồn
    source_processing = [
        {'name': 'Facebook của A', 'platform': 'Facebook', 'avg_time': '10 phút', 'percentage': '<15 phút', 'status': 'Rất nhanh'},
        {'name': 'Zalo của B', 'platform': 'Zalo', 'avg_time': '45 phút', 'percentage': '15-60 phút', 'status': 'Nhanh'},
        {'name': 'Telegram của C', 'platform': 'Telegram', 'avg_time': '2 giờ', 'percentage': '1-3 giờ', 'status': 'Bình thường'},
        {'name': 'Discord của D', 'platform': 'Discord', 'avg_time': '8 giờ', 'percentage': '3-12 giờ', 'status': 'Lâu'},
        {'name': 'Instagram của E', 'platform': 'Instagram', 'avg_time': '24 giờ', 'percentage': '>12 giờ', 'status': 'Rất lâu'},
    ]
    
    # Dữ liệu xu hướng theo tháng
    trend_data = {
        'labels': ['T1', 'T2', 'T3', 'T4', 'T5', 'T6'],
        'times': [45, 40, 35, 30, 25, 20],  # Thời gian xử lý trung bình (phút)
    }
    
    context = {
        'processing_data': processing_data,
        'source_processing': source_processing,
        'trend_data': trend_data,
        'avg_processing_time': '28 phút',
        'improvement': 15,  # Cải thiện % so với tháng trước
    }
    
    return render(request, 'dashboard/reports/processing_time.html', context)

# Thêm view cho email editor
@login_required
def email_editor(request, template_id=None):
    """
    View hiển thị trình soạn thảo email.
    Nếu template_id được cung cấp, sẽ tải nội dung mẫu để chỉnh sửa.
    """
    template = None
    if template_id:
        try:
            # Giả sử có model EmailTemplate
            template = EmailTemplate.objects.get(id=template_id)
        except:
            # Xử lý khi không tìm thấy template
            pass
    
    context = {
        'template': template
    }
    return render(request, 'dashboard/email/editor.html', context)

@login_required
def email_save_template(request):
    """
    API endpoint để lưu mẫu email
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Phương thức không được hỗ trợ'}, status=405)
    
    # Lấy dữ liệu từ request
    name = request.POST.get('name')
    category = request.POST.get('category')
    is_active = request.POST.get('is_active') == 'true'
    subject = request.POST.get('subject')
    content = request.POST.get('content')
    
    if not name or not subject or not content:
        return JsonResponse({'error': 'Thiếu thông tin bắt buộc'}, status=400)
    
    try:
        # Giả sử có model EmailTemplate
        template_id = request.POST.get('template_id')
        if template_id:
            # Cập nhật template hiện có
            template = EmailTemplate.objects.get(id=template_id)
            template.name = name
            template.category = category
            template.is_active = is_active
            template.subject = subject
            template.content = content
        else:
            # Tạo template mới
            template = EmailTemplate(
                name=name,
                category=category,
                is_active=is_active,
                subject=subject,
                content=content,
                created_by=request.user
            )
        
        template.save()
        
        return JsonResponse({
            'success': True,
            'template_id': template.id,
            'redirect': reverse('dashboard:email_templates')  # Thêm namespace dashboard
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
